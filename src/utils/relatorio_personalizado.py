import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_00, FORMAT_DATE_YYYYMMDD2
import os
from datetime import datetime
from io import BytesIO
from calendar import monthrange

def _to_first_day_of_month_str(data):
    """
    Recebe data (datetime.date, datetime.datetime, ou string) e retorna
    string no formato '01/MM/YYYY'.
    """
    if data is None:
        return ""
    # Se já for datetime/date
    if hasattr(data, "year") and hasattr(data, "month"):
        year = data.year
        month = data.month
    else:
        # Tentar parsear string em vários formatos comuns
        data_str = str(data)
        parsed = None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m", "%m/%Y", "%d-%m-%Y"):
            try:
                parsed = datetime.strptime(data_str, fmt)
                break
            except Exception:
                continue
        if parsed is None:
            # fallback: tentar extrair ano e mês via split
            try:
                parts = data_str.replace("/", "-").split("-")
                if len(parts) >= 2:
                    # assumir formatos como YYYY-MM ou DD-MM-YYYY
                    if len(parts[0]) == 4:
                        year = int(parts[0]); month = int(parts[1])
                    else:
                        year = int(parts[-1]); month = int(parts[-2])
                else:
                    # se falhar, usar data atual
                    now = datetime.now()
                    year, month = now.year, now.month
            except Exception:
                now = datetime.now()
                year, month = now.year, now.month
        else:
            year = parsed.year
            month = parsed.month

    return f"01/{month:02d}/{year}"

def gerar_relatorio_mensal_personalizado(registros, funcionarios, projetos, mes_ano=None):
    """
    Gera um relatório mensal personalizado (cada registro em linha separada).
    - Mês no formato '01/MM/YYYY' (primeiro dia do mês)
    - Horas Mês: total de horas do funcionário no mês correspondente ao registro
    - Horas Trabalhadas: horas do próprio registro (não agregado)
    - Observações por registro: colocadas na coluna adequada (9014/9010/9021)
    """
    # Criar um novo workbook e selecionar a planilha ativa
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planilha1"

    # Definir estilos
    header_font = Font(name='Calibri', size=11, bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    middle_border = Border(
        left=Side(style=None),
        right=Side(style=None),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Definir cabeçalhos na linha 3
    headers = [
        'COD Colaborador', 'Nome Colaborador', 'Mês', 'Horas Mês',
        'GP', 'Horas Trabalhadas', 'Proporção de Hora', 'Valor por GP',
        'Observação GP 9014', 'Observação GP 9010', 'Observação GP 9021'
    ]

    # Adicionar cabeçalhos começando na coluna C
    for i, header in enumerate(headers):
        col = i + 3  # Começando na coluna C (índice 3)
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        # borda superior
        cell.border = thin_border

    # --- Preparar cálculo de Horas Mês por funcionário por mês ---
    # Mapea (funcionario_id, YYYY-MM) -> total horas no mês
    horas_mes_por_funcionario_mes = {}
    for r in registros:
        # determinar key_mes a partir de r.data
        # tentamos extrair YYYY-MM
        if hasattr(r, "data") and r.data:
            d = r.data
            if hasattr(d, "year") and hasattr(d, "month"):
                key_mes = f"{d.year:04d}-{d.month:02d}"
            else:
                # tentar parse string
                try:
                    parsed = None
                    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m", "%m/%Y", "%d-%m-%Y"):
                        try:
                            parsed = datetime.strptime(str(d), fmt)
                            break
                        except Exception:
                            continue
                    if parsed:
                        key_mes = f"{parsed.year:04d}-{parsed.month:02d}"
                    else:
                        # fallback usar mes_ano passado ou data atual
                        if mes_ano:
                            key_mes = mes_ano
                        else:
                            now = datetime.now()
                            key_mes = f"{now.year:04d}-{now.month:02d}"
                except Exception:
                    if mes_ano:
                        key_mes = mes_ano
                    else:
                        now = datetime.now()
                        key_mes = f"{now.year:04d}-{now.month:02d}"
        else:
            key_mes = mes_ano if mes_ano else datetime.now().strftime("%Y-%m")

        func_id = getattr(r, "funcionario_id", None)
        horas = float(getattr(r, "horas_trabalhadas", 0) or 0)
        dict_key = (func_id, key_mes)
        horas_mes_por_funcionario_mes[dict_key] = horas_mes_por_funcionario_mes.get(dict_key, 0) + horas

    # --- Preencher dados: uma linha por registro ---
    row_index = 4  # Começar na linha 4 (após os cabeçalhos)

    # Helper para achar funcionário e projeto rapidamente por id
    funcionarios_map = {f.id: f for f in funcionarios}
    projetos_map = {p.id: p for p in projetos}

    for registro in registros:
        funcionario = funcionarios_map.get(registro.funcionario_id)
        projeto = projetos_map.get(registro.projeto_id)

        if not funcionario or not projeto:
            # opcional: pular se faltar informação essencial
            continue

        # Determinar mês do registro e string '01/MM/YYYY'
        mes_str = _to_first_day_of_month_str(getattr(registro, "data", None))

        # Determinar chave para Horas Mês (mes no formato YYYY-MM)
        d = getattr(registro, "data", None)
        if hasattr(d, "year") and hasattr(d, "month"):
            key_mes = f"{d.year:04d}-{d.month:02d}"
        else:
            # tentar parse ou usar mes_ano param
            try:
                parsed = None
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m", "%m/%Y", "%d-%m-%Y"):
                    try:
                        parsed = datetime.strptime(str(d), fmt)
                        break
                    except Exception:
                        continue
                if parsed:
                    key_mes = f"{parsed.year:04d}-{parsed.month:02d}"
                else:
                    key_mes = mes_ano if mes_ano else datetime.now().strftime("%Y-%m")
            except Exception:
                key_mes = mes_ano if mes_ano else datetime.now().strftime("%Y-%m")

        horas_mes_total = horas_mes_por_funcionario_mes.get((registro.funcionario_id, key_mes), 0.0)
        horas_trabalhadas_registro = float(getattr(registro, "horas_trabalhadas", 0) or 0)

        proporcao = (horas_trabalhadas_registro / horas_mes_total) if horas_mes_total > 0 else 0.0

        column_start = 3  # Coluna C

        # Preencher células
        ws.cell(row=row_index, column=column_start, value=funcionario.cod_funcionario)  # COD Colaborador
        ws.cell(row=row_index, column=column_start+1, value=funcionario.nome)  # Nome Colaborador
        # Mês = primeiro dia do mês no formato dd/mm/YYYY
        cell_mes = ws.cell(row=row_index, column=column_start+2, value=mes_str)  # Mês
        # Forçar formato de texto (já está como string '01/MM/YYYY'); se quiser como data real:
        # parse '01/MM/YYYY' e setar value como datetime e number_format = 'DD/MM/YYYY'
        try:
            dt_for_cell = datetime.strptime(mes_str, "%d/%m/%Y")
            cell_mes.value = dt_for_cell
            cell_mes.number_format = "DD/MM/YYYY"
        except Exception:
            pass

        # Horas Mês (total do funcionário no mês do registro)
        cell_horas_mes = ws.cell(row=row_index, column=column_start+3, value=horas_mes_total)
        cell_horas_mes.number_format = '0.00'

        # GP (código do projeto)
        ws.cell(row=row_index, column=column_start+4, value=str(projeto.cod))

        # Horas Trabalhadas (apenas do registro)
        cell_horas_trab = ws.cell(row=row_index, column=column_start+5, value=horas_trabalhadas_registro)
        cell_horas_trab.number_format = '0.00'

        # Proporção de Hora
        cell_proporcao = ws.cell(row=row_index, column=column_start+6, value=proporcao)
        cell_proporcao.number_format = '0.00%'

        # Valor por GP (mantido em branco aqui, conforme template)
        # ws.cell(row=row_index, column=column_start+7, value="")

        # Observações: preencher somente a coluna correspondente ao GP deste registro
        obs = getattr(registro, "observacoes", "") or ""
        if str(projeto.cod) == "9014":
            ws.cell(row=row_index, column=column_start+8, value=obs)
        elif str(projeto.cod) == "9010":
            ws.cell(row=row_index, column=column_start+9, value=obs)
        elif str(projeto.cod) == "9021":
            ws.cell(row=row_index, column=column_start+10, value=obs)
        # caso seja outro projeto, não preenche nenhuma das colunas de observação

        # Aplicar bordas às células de dados (Colunas C a M -> 3 a 13)
        for col in range(3, 14):  # Colunas C a M
            cell = ws.cell(row=row_index, column=col)
            # aplicar borda inferior fina
            cell.border = Border(bottom=Side(style='thin'))

        row_index += 1

    # Ajustar largura das colunas
    for col in range(3, 14):  # Colunas C a N
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 17

    # Coluna de nome do colaborador mais larga (coluna D = index 4)
    ws.column_dimensions['D'].width = 25

    # Salvar o arquivo em memória
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output

def adaptar_exportacao_relatorio_mensal(db, funcionario_id=None, projeto_id=None, mes_ano=None):
    """
    Adaptador para exportação no formato personalizado.
    Retorna (arquivo_bytes_em_memoria, nome_arquivo).
    """
    try:
        # Obter os registros filtrados do DB
        registros = db.listar_registros_horas(
            funcionario_id=funcionario_id,
            projeto_id=projeto_id,
            mes_ano=mes_ano
        )

        # Listas de funcionários e projetos
        funcionarios = db.listar_usuarios()
        # remover administradores caso necessário
        funcionarios = [f for f in funcionarios if getattr(f, "tipo", None) != "administrador"]
        projetos = db.listar_projetos()

        if not registros:
            raise ValueError("Nenhum registro encontrado para exportação")

        # nome do arquivo (usar mes_ano se fornecido)
        mes_ano_str = mes_ano if mes_ano else datetime.now().strftime("%Y-%m")
        nome_arquivo = f"controle_horas_{mes_ano_str}.xlsx"

        # Gerar o arquivo em memória
        arquivo_bytes = gerar_relatorio_mensal_personalizado(
            registros=registros,
            funcionarios=funcionarios,
            projetos=projetos,
            mes_ano=mes_ano
        )

        return arquivo_bytes, nome_arquivo

    except Exception as e:
        # Log simples para debugging; re-levanta exceção para a camada superior tratar
        print(f"[adaptar_exportacao_relatorio_mensal] Erro: {e}")
        raise
